#!/usr/bin/env python
# -*- coding: utf-8 -*-
import time

import openpyxl
import sys
import os
from openpyxl.styles import Alignment
import json
import codecs
import re

curPath = os.path.abspath(os.path.dirname(__file__))
rootPath = curPath[:curPath.find("Interface_Auto_Test\\") + len("Interface_Auto_Test\\")]
sys.path.append(rootPath)

class Read_xls:

	def __init__(self, path, sheetname):
		self.path = path
		# 打开文件，获取excel文件的workbook（工作簿）对象
		# self.workbook = xlrd.open_workbook(self.path, formatting_info=True)
		self.workbook = openpyxl.load_workbook(self.path)
		self.sheetname = sheetname

	def read(self):

		# print(path)
		# 打开excel文件,open_workbook(path),path为excel所在的路径
		# 打开excel表,这里表示打开第一张表
		table = self.workbook.get_sheet_by_name(self.sheetname)

		# nrows = table.max_row		# 获取excel的行数
		# print(nrows)
		# ncols = table.max_column		#获取excel的列数
		# print(ncols)
		# keys = table.row_values(0)		#获取第一行的值, 返回一个list
		# print(keys)

		rows = []
		for row in table.iter_rows():
			rows.append(row)

		cols = []
		for col in table.iter_cols():
			cols.append(col)

		resp = []		#创建一个list，用于存放
		x = 1
		for i in range(len(rows)-1):
			s = {}
			# print(i)
			s['rowNum'] = i+2 	#加入用例的行数，用户后面写入数据
			values = rows[x]	# 获取每一行的值
			# print(values)
			for j in range(len(cols)):
				# print('j=',j)
				s[rows[0][j].value] = values[j].value
			# print(s)
			resp.append(s)
			x += 1

		return resp

	def write(self, testdata, resp, resp_time):
		table = self.workbook.get_sheet_by_name(self.sheetname)

		# 写入预期结果, 下标从1开始算
		start = time.time()
		# table.cell(row=testdata['rowNum'], column=12, value=str(resp))
		# 设置单元格格式
		align = Alignment(horizontal='left', vertical='center', wrap_text=True)
		# 写入response
		table.cell(row=testdata['rowNum'], column=12, value=str(resp)).alignment = align
		# 写入响应时间
		table.cell(row=testdata['rowNum'], column=13, value=str(resp_time))

		self.workbook.save(rootPath + r'docs/data_copy.xlsx')
		print(time.time() - start, "33333333333333333333333333333333333333333333333")

	# 转换成json文件, 供postman使用
	def get_dict(self, resp, rootPath=None):
		list1 = []
		for res in resp:
			# 转换body
			# body = res['body'].replace(r"\n", "").replace(" ", "").replace(r'"', r'\"')
			# print(body)
			pathlist = "".join(re.findall("[^8080]+(?!.*/)", res['URL'])).split("/")
			pathlist.remove('')


			dict1 = {
				"name": res['name'],
				"event": [
					{
						"listen": "test",
						"script": {
							"id": "ebcbfafb-e5a1-479a-ad0d-afbcf2501538",
							"exec": [
								"pm.test(\"Status code is 200\", function () {",
								"    pm.response.to.have.status(200);",
								"});"
							],
							"type": "text/javascript"
						}
					},
				],
				"request": {
					"method": res['method'],
					"header": [
						{
							"key": "Content-Type",
							"value": "application/x-www-form-urlencoded"
						},
						{
							"key": "User-Agent",
							"value": "Openwave"
						},
						{
							"key": "Content-Language",
							"value": "en-US"
						}
					],
					"body": {
						"mode": "formdata",
						"formdata": [
							{
								"key": "QueryString",
								"value": res['body'],
								"type": "text"
							}
						]
					},
					"url": {
						"raw": res['URL'],
						"protocol": "http",
						"host": [
							"trade-route-develop",
							"ntdev",
							"be"
						],
						"port": "8080",
						"path": pathlist
					}
				},
				"response": []
			}

			list1.append(dict1)

		# print(list1)

		head_dict = {
			"info": {
				"_postman_id": "63ef4c63-f578-46c3-9888-6c762a7ba3ec",
				"name": self.sheetname,
				"schema": "https://schema.getpostman.com/json/collection/v2.1.0/collection.json"
			},
			"protocolProfileBehavior": {}
		}
		head_dict['item'] = list1
		# print(head_dict)
		print(json.dumps(head_dict, ensure_ascii=False))

		# 保存为utf-8编码的文件
		filename = rootPath + r"docs/{}.json".format(self.sheetname)
		try:
		    f=codecs.open(filename, 'w', 'UTF-8')
		    f.write(json.dumps(head_dict, ensure_ascii=False))
		    f.close()
		except Exception as e:
		    print(e)
		os.system('pause')



if __name__ == "__main__":
	# path = 'E:/郑某人/Python_Demo/Interface_Test_Frame/data/data.xls'
	curPath = os.path.abspath(os.path.dirname(__file__))
	rootPath = curPath[:curPath.find("Interface_Auto_Test\\") + len("Interface_Auto_Test\\")]
	path = rootPath+r'docs/data.xlsx'
	reads = Read_xls(path, u"委托下单")
	# reads.write({"rowNum" : 3}, "2222222")
	resp = reads.read()
	# print(resp)
	reads.get_dict(resp, rootPath)


