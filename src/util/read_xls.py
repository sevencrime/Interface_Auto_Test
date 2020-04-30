#!/usr/bin/env python
# -*- coding: utf-8 -*-
import time

import openpyxl
import sys
import os
from openpyxl.styles import Alignment
import json
import codecs
import re

curPath = os.path.abspath(os.path.dirname(__file__))
rootPath = curPath[:curPath.find("Interface_Auto_Test\\") + len("Interface_Auto_Test\\")]
sys.path.append(rootPath)


class Read_xls:

    def __init__(self, path, sheetname):
        self.path = path
        # 打开文件，获取excel文件的workbook（工作簿）对象
        # self.workbook = xlrd.open_workbook(self.path, formatting_info=True)
        self.workbook = openpyxl.load_workbook(self.path)
        self.sheetname = sheetname

    def read(self):

        keyName = {
            "用例编号": "case",
            "功能模块": "module",
            "子模块": "Submodule",
            "用例等级": "level",
            "测试类别": "test_type",
            "用例标题": "name",
            "预置条件": "Preconditions",
            "接口地址": "URL",
            "请求方式": "method",
            "请求头": "headers",
            "请求参数": "body",
            "期望结果": "Expected_Response",
            "响应时间": "Resphone_Time",
            "测试结果": "result",
            "测试轮次": "Test_Round",
            "执行人": "tester",
            "备注": "remark",
            "操作步骤": "step",
            "预期结果": "expected",
            "操作步骤": "step"
        }

        # 打开excel表,这里表示打开第一张表
        table = self.workbook.get_sheet_by_name(self.sheetname)

        rows = []
        for row in table.iter_rows():
            rows.append(row)

        cols = []
        for col in table.iter_cols():
            cols.append(col)

        resp = []  # 创建一个list，用于存放
        x = 1
        for i in range(len(rows) - 1):
            s = {}
            # print(i)
            s['rowNum'] = i + 2  # 加入用例的行数，用户后面写入数据
            values = rows[x]  # 获取每一行的值
            # print(values)
            for j in range(len(cols)):
                # print('j=',j)
                s[keyName[rows[0][j].value]] = values[j].value
            # print(s)
            resp.append(s)
            x += 1

        return resp

    def write(self, testdata, resp, resp_time):
        table = self.workbook.get_sheet_by_name(self.sheetname)

        # 写入预期结果, 下标从1开始算
        start = time.time()
        # table.cell(row=testdata['rowNum'], column=12, value=str(resp))
        # 设置单元格格式
        align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        # 写入response
        table.cell(row=testdata['rowNum'], column=12, value=str(resp)).alignment = align
        # 写入响应时间
        table.cell(row=testdata['rowNum'], column=13, value=str(resp_time))

        self.workbook.save(rootPath + r'docs/data_copy.xlsx')
        print(time.time() - start, "33333333333333333333333333333333333333333333333")

    # 转换成json文件, 供postman使用
    def get_dict(self, resp, rootPath=None):
        list1 = []
        for res in resp:
            # 转换body
            # body = res['body'].replace(r"\n", "").replace(" ", "").replace(r'"', r'\"')
            # print(body)
            pathlist = "".join(re.findall("[^8080]+(?!.*/)", res['URL'])).split("/")
            pathlist.remove('')

            dict1 = {
                "name": res['name'],
                "event": [
                    {
                        "listen": "test",
                        "script": {
                            "id": "ebcbfafb-e5a1-479a-ad0d-afbcf2501538",
                            "exec": [
                                "pm.test(\"Status code is 200\", function () {",
                                "    pm.response.to.have.status(200);",
                                "});"
                            ],
                            "type": "text/javascript"
                        }
                    },
                ],
                "request": {
                    "method": res['method'],
                    "header": [
                        {
                            "key": "Content-Type",
                            "value": "application/x-www-form-urlencoded"
                        },
                        {
                            "key": "User-Agent",
                            "value": "Openwave"
                        },
                        {
                            "key": "Content-Language",
                            "value": "en-US"
                        }
                    ],
                    "body": {
                        "mode": "formdata",
                        "formdata": [
                            {
                                "key": "QueryString",
                                "value": res['body'],
                                "type": "text"
                            }
                        ]
                    },
                    "url": {
                        "raw": res['URL'],
                        "protocol": "http",
                        "host": [
                            "trade-route-develop",
                            "ntdev",
                            "be"
                        ],
                        "port": "8080",
                        "path": pathlist
                    }
                },
                "response": []
            }

            list1.append(dict1)

        # print(list1)

        head_dict = {
            "info": {
                "_postman_id": "63ef4c63-f578-46c3-9888-6c762a7ba3ec",
                "name": self.sheetname,
                "schema": "https://schema.getpostman.com/json/collection/v2.1.0/collection.json"
            },
            "protocolProfileBehavior": {}
        }
        head_dict['item'] = list1
        # print(head_dict)
        print(json.dumps(head_dict, ensure_ascii=False))

        # 保存为utf-8编码的文件
        filename = rootPath + r"docs/{}.json".format(self.sheetname)
        try:
            f = codecs.open(filename, 'w', 'UTF-8')
            f.write(json.dumps(head_dict, ensure_ascii=False))
            f.close()
        except Exception as e:
            print(e)
        os.system('pause')

    # 转换成json文件, 供postman使用
    def temp_get_dict(self, resp, rootPath=None, **kwargs):

        list1 = []
        for res in resp:
            name = res['name']
            method = kwargs.get("method")
            body = res['step']
            uri = kwargs.get("uri")
            headers = kwargs.get("headers")

            # 转换body
            body = re.search(r'[{](.*?)[}]', body.replace("\n", "").replace(" ", "")).group()
            pathlist = "".join(re.findall("[^8080]+(?!.*/)", uri)).split("/")
            pathlist.remove('{{server}}')


            # 构造body中x-www-form-urlencoded
            bodylist = []
            for k, v in eval(body).items():
                bitem = {}
                bitem["key"] = k
                bitem["value"] = v
                bitem["type"] = "text"
                bodylist.append(bitem)

            dict1 = {
                "name": name,
                "request": {
                    "method": method,
                    "header": headers,
                    "body": {
                        "mode": "urlencoded",
                        "urlencoded": bodylist
                    },
                    "url": {
                        "raw": uri,
                        "host": [
                            "{{server}}"
                        ],
                        "path": pathlist
                    }
                },
                "response": []
            }
            list1.append(dict1)

        # print(list1)

        head_dict = {
            "info": {
                "_postman_id": "63ef4c63-f578-46c3-9888-6c762a7ba3ec",
                "name": self.sheetname,
                "schema": "https://schema.getpostman.com/json/collection/v2.1.0/collection.json"
            },
            "protocolProfileBehavior": {}
        }
        head_dict['item'] = list1
        # print(head_dict)
        print(json.dumps(head_dict, ensure_ascii=False))

        # 保存为utf-8编码的文件
        filename = rootPath + r"docs/{}.json".format(self.sheetname)
        try:
            f = codecs.open(filename, 'w', 'UTF-8')
            f.write(json.dumps(head_dict, ensure_ascii=False))
            f.close()
        except Exception as e:
            print(e)
        os.system('pause')


def run_Ayers_json(rootPath):
    path = rootPath + r'docs/ayers网关接口测试用例.xlsx'
    reads = Read_xls(path, u"委托撤单")



# 统一认证
def run_certification_json(rootPath, sheetname):
    path = rootPath + r'docs/统一认证中心.xlsx'
    reads = Read_xls(path, sheetname)
    # reads.write({"rowNum" : 3}, "2222222")
    resp = reads.read()
    # print(resp)
    # reads.get_dict(resp, rootPath)
    uri = "{{server}}/v2/inward/users"
    headers = [
        {
            "key": "x-api-key",
            "value": "cm9vdDphZG1pbg==",
            "type": "text"
        },
        {
            "key": "Authorization",
            "value": "Basic {{Authorization}}",
            "type": "text"
        }
    ]
    reads.temp_get_dict(resp, rootPath, method="post", uri=uri, headers=headers)



def get_jmx(rootPath, sheetname, **kwargs):

    path = rootPath + r'docs/统一认证中心.xlsx'
    reads = Read_xls(path, sheetname)
    # reads.write({"rowNum" : 3}, "2222222")
    resp = reads.read()


    jmxstr = '''<?xml version="1.0" encoding="UTF-8"?>
<jmeterTestPlan version="1.2" properties="5.0" jmeter="5.0 r1840935">
  <hashTree>
    <TestPlan guiclass="TestPlanGui" testclass="TestPlan" testname="测试计划" enabled="true">
      <stringProp name="TestPlan.comments"></stringProp>
      <boolProp name="TestPlan.functional_mode">false</boolProp>
      <boolProp name="TestPlan.tearDown_on_shutdown">true</boolProp>
      <boolProp name="TestPlan.serialize_threadgroups">false</boolProp>
      <elementProp name="TestPlan.user_defined_variables" elementType="Arguments" guiclass="ArgumentsPanel" testclass="Arguments" testname="用户定义的变量" enabled="true">
        <collectionProp name="Arguments.arguments"/>
      </elementProp>
      <stringProp name="TestPlan.user_define_classpath"></stringProp>
    </TestPlan>
    <hashTree>
      <HeaderManager guiclass="HeaderPanel" testclass="HeaderManager" testname="HTTP信息头管理器" enabled="true">
        <collectionProp name="HeaderManager.headers">
          <elementProp name="" elementType="Header">
            <stringProp name="Header.name">x-api-key</stringProp>
            <stringProp name="Header.value">cm9vdDphZG1pbg==</stringProp>
          </elementProp>
          <elementProp name="" elementType="Header">
            <stringProp name="Header.name">Authorization</stringProp>
            <stringProp name="Header.value">Basic dGVzdGFwcDI6YWJjZA==</stringProp>
          </elementProp>
        </collectionProp>
      </HeaderManager>
      <hashTree/>
      <ThreadGroup guiclass="ThreadGroupGui" testclass="ThreadGroup" testname='{sheetname}' enabled="true">
        <stringProp name="ThreadGroup.on_sample_error">continue</stringProp>
        <elementProp name="ThreadGroup.main_controller" elementType="LoopController" guiclass="LoopControlPanel" testclass="LoopController" testname="循环控制器" enabled="true">
          <boolProp name="LoopController.continue_forever">false</boolProp>
          <stringProp name="LoopController.loops">1</stringProp>
        </elementProp>
        <stringProp name="ThreadGroup.num_threads">1</stringProp>
        <stringProp name="ThreadGroup.ramp_time">1</stringProp>
        <boolProp name="ThreadGroup.scheduler">false</boolProp>
        <stringProp name="ThreadGroup.duration"></stringProp>
        <stringProp name="ThreadGroup.delay"></stringProp>
      </ThreadGroup>
      <hashTree>
        {httplist}
      </hashTree>
    </hashTree>
  </hashTree>
</jmeterTestPlan>
    '''


    httpstr = '''
<HTTPSamplerProxy guiclass="HttpTestSampleGui" testclass="HTTPSamplerProxy" testname='{casename}' enabled="true">
  <elementProp name="HTTPsampler.Arguments" elementType="Arguments" guiclass="HTTPArgumentsPanel" testclass="Arguments" testname="用户定义的变量" enabled="true">
    <collectionProp name="Arguments.arguments">
      {paramslist}
    </collectionProp>
  </elementProp>
  <stringProp name="HTTPSampler.domain">{domain}</stringProp>
  <stringProp name="HTTPSampler.port">{port}</stringProp>
  <stringProp name="HTTPSampler.protocol">{protocol}</stringProp>
  <stringProp name="HTTPSampler.contentEncoding"></stringProp>
  <stringProp name="HTTPSampler.path">{path}</stringProp>
  <stringProp name="HTTPSampler.method">{method}</stringProp>
  <boolProp name="HTTPSampler.follow_redirects">true</boolProp>
  <boolProp name="HTTPSampler.auto_redirects">false</boolProp>
  <boolProp name="HTTPSampler.use_keepalive">true</boolProp>
  <boolProp name="HTTPSampler.DO_MULTIPART_POST">false</boolProp>
  <stringProp name="HTTPSampler.embedded_url_re"></stringProp>
  <stringProp name="HTTPSampler.connect_timeout"></stringProp>
  <stringProp name="HTTPSampler.response_timeout"></stringProp>
</HTTPSamplerProxy>
<hashTree/>
    '''

    params = '''
<elementProp name="phone_number" elementType="HTTPArgument">
  <boolProp name="HTTPArgument.always_encode">false</boolProp>
  <stringProp name="Argument.value">{value}</stringProp>
  <stringProp name="Argument.metadata">=</stringProp>
  <boolProp name="HTTPArgument.use_equals">true</boolProp>
  <stringProp name="Argument.name">{key}</stringProp>
</elementProp>
    '''

    httplist = []

    for res in resp:
        name = res['name']
        method = kwargs.get("method")
        body = res['step']
        uri = kwargs.get("uri")
        headers = kwargs.get("headers")

        # 转换body
        body = re.search(r'[{](.*?)[}]', body.replace("\n", "").replace(" ", "")).group()

        paramlist = []
        for k, v in eval(body).items():
            param = params.format(key=k, value=v)
            paramlist.append(param)


        http_format = httpstr.format(
                    casename=name,
                    paramslist="".join(paramlist),
                    domain=uri,
                    port=kwargs.get("port"),
                    protocol=kwargs.get("protocol"),
                    method=method,
                    path=kwargs.get("path")
                )


        httplist.append(http_format)

    jmxstr = jmxstr.format(sheetname=sheetname, httplist="".join(httplist))
    # print(jmxstr)

    # 保存为utf-8编码的文件
    filename = rootPath + r"docs/{}.jmx".format(sheetname)
    try:
        f = codecs.open(filename, 'w', 'UTF-8')
        f.write(jmxstr)
        f.close()
    except Exception as e:
        print(e)
    os.system('pause')



if __name__ == "__main__":
    # path = 'E:/郑某人/Python_Demo/Interface_Test_Frame/data/data.xls'
    curPath = os.path.abspath(os.path.dirname(__file__))
    rootPath = curPath[:curPath.find("Interface_Auto_Test\\") + len("Interface_Auto_Test\\")]
    # path = rootPath + r'docs/ayers网关接口测试用例.xlsx'
    # reads = Read_xls(path, u"委托撤单")

    # run_certification_json(rootPath, u"V1-密码方式获取token")
    uri = "eddid-auth-center-develop.eddidone.be"
    port = "180"
    protocol = "http"
    path="/v2/inward/users"
    method = "post"
    get_jmx(rootPath, u"注册", method=method.upper(), uri=uri, port=port, protocol=protocol, path=path)